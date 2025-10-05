# files/api_views.py
from datetime import timedelta, date, datetime
from django.utils import timezone
from django.db.models import Count
from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse

from .models import Document, Category
from .serializers import DocumentSerializer, CategorySerializer

# ======= CRUD عبر DRF Router =======
class DocumentViewSet(viewsets.ModelViewSet):
    queryset = Document.objects.all().select_related("category")
    serializer_class = DocumentSerializer
    permission_classes = [IsAuthenticated]


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]


# ======= Dashboard JSON =======
class FilesDashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        today = timezone.localdate()
        docs = Document.objects.select_related("category")

        total = docs.count()
        expired = docs.filter(expires_on__lt=today).count()
        expiring_7 = docs.filter(
            expires_on__gte=today,
            expires_on__lte=today + timedelta(days=7)
        ).count()
        active = total - expired

        by_category = (
            docs.values("category__name", "category__color")
            .annotate(count=Count("id"))
            .order_by("category__name")
        )

        upcoming = []
        for d in docs.filter(expires_on__gte=today).order_by("expires_on")[:6]:
            upcoming.append({
                "id": d.id,
                "title": d.title,
                "expires_on": d.expires_on,
                "days_to_expiry": d.days_to_expiry,
                "category": (
                    {
                        "name": d.category.name,
                        "color": d.category.color,
                    } if d.category else None
                ),
            })

        status_json = {
            "labels": ["Active", "≤ 7 days", "Expired"],
            "values": [active, expiring_7, expired],
        }
        cat_json = {
            "labels": [c["category__name"] or "Uncategorized" for c in by_category],
            "values": [c["count"] for c in by_category],
        }
        up_json = {
            "labels": [d["title"] for d in upcoming],
            "values": [d["days_to_expiry"] for d in upcoming],
        }

        return Response({
            "total": total,
            "active": active,
            "expiring_7": expiring_7,
            "expired": expired,
            "by_category": list(by_category),
            "upcoming": upcoming,
            "status_json": status_json,
            "cat_json": cat_json,
            "up_json": up_json,
        })


# ======= Export / Template (ترجع ملفات للتحميل) =======
from openpyxl import Workbook, load_workbook
import csv
import io

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_csv_api(request):
    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = 'attachment; filename="documents.csv"'

    writer = csv.writer(resp)
    writer.writerow(["Title", "Category", "Expires On",
                     "Owner Email", "Alert Window Days", "Notes"])

    qs = Document.objects.select_related('category').order_by('expires_on', 'title')
    for d in qs:
        writer.writerow([
            d.title,
            d.category.name if d.category else '',
            d.expires_on.isoformat() if d.expires_on else '',
            d.owner_email or '',
            d.alert_window_days,
            (d.notes or '').replace('\r', ' ').replace('\n', ' ').strip(),
        ])
    return resp


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_xlsx_api(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"
    ws.append(["Title", "Category", "Expires On",
               "Owner Email", "Alert Window Days", "Notes"])

    qs = Document.objects.select_related('category').order_by('expires_on', 'title')
    for d in qs:
        ws.append([
            d.title,
            d.category.name if d.category else '',
            d.expires_on.isoformat() if d.expires_on else '',
            d.owner_email or '',
            d.alert_window_days,
            d.notes or '',
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp['Content-Disposition'] = 'attachment; filename="documents.xlsx"'
    return resp


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def import_template_api(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"
    ws.append(["Title", "Category", "Expires On",
               "Owner Email", "Alert Window Days", "Notes"])
    ws.append([
        "AHMED KAMEL", "VISA EMP", "2025-10-05",
        "kamel.me311@gmail.com", 7,
        "UID: 7841..., Policy: C/01/2023/..."
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp['Content-Disposition'] = 'attachment; filename=\"expiryvault_template.xlsx\"'
    return resp


# ======= Import (.xlsx) =======
from rest_framework.parsers import MultiPartParser, FormParser, FileUploadParser

@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser, FileUploadParser])
def import_xlsx_api(request):
    """
    Import documents from uploaded Excel file.
    Expected headers (case/space-insensitive): 
    ["Title", "Category", "Expires On", "Owner Email", "Alert Window Days", "Notes"]
    """
    # احصل على الملف من multipart أو من data (في حال FileUploadParser)
    file_obj = request.FILES.get("file") or request.data.get("file")
    if not file_obj:
        return Response(
            {"error": "No file uploaded. Send as multipart/form-data with field name 'file'."},
            status=status.HTTP_400_BAD_REQUEST
        )

    # تحقق الامتداد اختياريًا
    fname = getattr(file_obj, "name", "") or ""
    if fname and not fname.lower().endswith(".xlsx"):
        return Response({"error": "Only .xlsx files are supported"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb = load_workbook(file_obj, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        if not rows or len(rows) < 2:
            return Response({"error": "File is empty or missing data"}, status=status.HTTP_400_BAD_REQUEST)

        # مطابقة مرنة للهيدر
        expected = ["Title", "Category", "Expires On", "Owner Email", "Alert Window Days", "Notes"]
        header = [(h or "").strip() for h in rows[0]]

        def norm(s):
            return (s or "").strip().lower()

        if [norm(h) for h in header] != [norm(h) for h in expected]:
            return Response(
                {"error": "Invalid file format.",
                 "expected": expected,
                 "got": header},
                status=status.HTTP_400_BAD_REQUEST
            )

        # استيراد الصفوف
        imported = 0
        from openpyxl.utils.datetime import from_excel as excel_to_dt

        for row in rows[1:]:
            if not row or not any(row):
                continue

            title, category_name, expires_on, owner_email, alert_days, notes = row

            # تحويل التاريخ إن كان رقم Excel أو نص ISO
            if isinstance(expires_on, (datetime, date)):
                # ok
                pass
            elif isinstance(expires_on, (int, float)):
                try:
                    expires_on = excel_to_dt(expires_on)
                except Exception:
                    expires_on = None
            elif isinstance(expires_on, str):
                try:
                    # يقبل "YYYY-MM-DD" أو "YYYY-MM-DDTHH:MM:SS"
                    expires_on = datetime.fromisoformat(expires_on)
                except Exception:
                    expires_on = None
            else:
                expires_on = None

            # حول datetime إلى date إن لزم
            if isinstance(expires_on, datetime):
                expires_on = expires_on.date()

            # alert_days إلى int
            try:
                alert_days = int(alert_days) if alert_days is not None else 0
            except Exception:
                alert_days = 0

            category = None
            if category_name:
                category, _ = Category.objects.get_or_create(name=str(category_name).strip())

            Document.objects.create(
                title=str(title).strip() if title else "",
                category=category,
                expires_on=expires_on,
                owner_email=(owner_email or "")[:254],
                alert_window_days=alert_days,
                notes=notes or "",
            )
            imported += 1

        return Response({"message": f"Imported {imported} documents"}, status=status.HTTP_200_OK)

    except Exception as e:
        # سجّل الستاك تريس لتسهيل الديبج
        import logging, traceback
        logging.getLogger(__name__).error("Import failed: %s\n%s", e, traceback.format_exc())
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
